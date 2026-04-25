import io, math, calendar
from datetime import date, timedelta
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_FILE  = "inventory_v2.xlsx"
TODAY       = date(2026, 4, 10)    # change to date.today() in production
ROUND_TO    = 500
COVER_DAYS  = 45

# ── Colors ────────────────────────────────────────────────────────────────────
C_BLUE   = "#1F4E79"
C_ORANGE = "#ED7D31"
C_GREEN  = "#375623"
C_RED    = "#C00000"
C_PURPLE = "#7030A0"
C_GRAY   = "#888888"
C_TEAL   = "#008080"

# ── Excel styles ──────────────────────────────────────────────────────────────
THIN = Border(
    left=Side(style="thin",color="D9D9D9"), right=Side(style="thin",color="D9D9D9"),
    top=Side(style="thin",color="D9D9D9"),  bottom=Side(style="thin",color="D9D9D9"))
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR2_FILL = PatternFill("solid", start_color="2E75B6")
ALT_FILL  = PatternFill("solid", start_color="EBF3FB")
WH_FILL   = PatternFill("solid", start_color="FFFFFF")
RED_FILL  = PatternFill("solid", start_color="FCE4D6")
YLW_FILL  = PatternFill("solid", start_color="FFF2CC")
GRN_FILL  = PatternFill("solid", start_color="E2EFDA")


def hcell(ws, row, col, val, w=None, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill      = fill if fill else HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = THIN
    if w: ws.column_dimensions[get_column_letter(col)].width = w
    return c


def dcell(ws, row, col, val, bold=False, color="000000", fill=None, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10, bold=bold, color=color)
    c.fill      = fill if fill else (ALT_FILL if row % 2 == 0 else WH_FILL)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = THIN
    return c


def sec(ws, row, col, val, n, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    c.fill      = fill if fill else HDR2_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = THIN
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+n-1)
    ws.row_dimensions[row].height = 20


# ── Utilities ─────────────────────────────────────────────────────────────────
def days_in_month(y, m): return calendar.monthrange(y, m)[1]
def round_up_to(v, u):   return math.ceil(max(v, 1) / u) * u


# ── Holt-Winters Multiplicative ───────────────────────────────────────────────

def holt_winters_mult(demand, multipliers, alpha, beta, gamma, phi):
    n   = len(demand)
    L   = np.empty(n)
    B   = np.empty(n)
    S   = np.array(multipliers, dtype=float).copy()  # initialize from col E

    # Initialize level from de-seasonalized first value
    deseas = demand / np.where(S > 0, S, 1.0)
    L[0]   = deseas[0]
    B[0]   = (deseas[1] - deseas[0]) if n > 1 else 0.0

    fitted    = np.empty(n)
    fitted[0] = L[0] * S[0]

    for t in range(1, n):
        s_t    = max(S[t], 0.01)
        L[t]   = alpha * (demand[t] / s_t) + (1 - alpha) * (L[t-1] + phi * B[t-1])
        B[t]   = beta  * (L[t] - L[t-1])  + (1 - beta)  * phi * B[t-1]
        S[t]   = gamma * (demand[t] / max(L[t], 0.01)) + (1 - gamma) * S[t]
        fitted[t] = L[t] * S[t]

    return L, B, S, fitted


def hw_forecast(L, B, future_multipliers, phi):
    """
    Holt-Winters h-step ahead forecast with damped trend.
    Future seasonal factors = user-provided future multipliers.

    F[t+h] = (L[t] + sum_{i=1}^{h} phi^i * B[t]) * future_mult[h]
    """
    out     = []
    phi_acc = 0.0
    for i, sm in enumerate(future_multipliers):
        phi_acc += phi ** (i + 1)
        level_proj = L[-1] + phi_acc * B[-1]
        out.append(round(level_proj * sm, 0))
    return np.array(out)


# ── Data I/O ──────────────────────────────────────────────────────────────────

def read_settings(wb):
    ws   = wb["Settings"]
    rows = list(ws.iter_rows(values_only=True))
    prod_cfg = {}; fore_cfg = {}; mode = None
    for row in rows:
        key = str(row[0] or "").strip(); kl = key.lower()
        if "inventory per product" in kl: mode = None; continue
        if kl == "product": mode = "products"; continue
        if "forecast & replenishment" in kl: mode = "forecast"; continue
        if mode == "products":
            if not key or row[1] is None: continue
            if any(k in kl for k in ["lead","alpha","beta","gamma","damp","fore","order","hold"]):
                mode = None; continue
            try:
                prod_cfg[key] = {"inventory": float(row[1] or 0),
                                 "rop":       float(row[2] or 0),
                                 "min_order": float(row[3] or 0)}
            except (TypeError, ValueError): continue
        elif mode == "forecast":
            if not key or row[1] is None: continue
            try: fore_cfg[key] = float(row[1])
            except (TypeError, ValueError): continue
    return prod_cfg, fore_cfg


def read_sheet(ws):
    hist = []
    for row in ws.iter_rows(min_row=4, max_row=16, values_only=True):
        if row[0] and not isinstance(row[0], (int, float)) and row[2] is not None:
            hist.append({"date":       pd.Timestamp(row[0]).date().replace(day=1),
                         "demand":     float(row[2]),
                         "multiplier": float(row[4]) if row[4] else 1.0})
    fut = []; in_f = False
    for row in ws.iter_rows(min_row=17, max_row=35, values_only=True):
        if row[0] and "future" in str(row[0]).lower() and row[1] and "future" in str(row[1]).lower():
            in_f = True; continue
        if in_f and row[0] and row[1] is not None:
            try:
                fut.append({"date":       pd.Timestamp(row[0]).date().replace(day=1),
                            "multiplier": float(row[1])})
            except (TypeError, ValueError): continue
    hist_df = pd.DataFrame(hist).sort_values("date").reset_index(drop=True)
    fut_df  = pd.DataFrame(fut).sort_values("date").reset_index(drop=True)
    return hist_df, fut_df


def run_forecast(hist_df, fut_df, alpha, beta, gamma, phi):
    demand      = hist_df["demand"].to_numpy(float)
    multipliers = hist_df["multiplier"].to_numpy(float)

    L, B, S, fitted = holt_winters_mult(demand, multipliers, alpha, beta, gamma, phi)

    last   = hist_df["date"].iloc[-1]
    future = fut_df[fut_df["date"] > last].copy().reset_index(drop=True)
    if future.empty:
        return pd.DataFrame(columns=["date","forecast","multiplier"]), L, B, S, fitted

    future["forecast"] = hw_forecast(L, B, future["multiplier"].to_numpy(), phi)
    return future[["date","forecast","multiplier"]], L, B, S, fitted


# ── Day-by-day simulation ─────────────────────────────────────────────────────

def simulate(hist_df, fore_df, cfg, lead):
    rop       = cfg["rop"]
    min_order = cfg["min_order"]
    stock     = cfg["inventory"]

    # Build daily rate map: current partial month + forecast months
    last   = hist_df.iloc[-1]
    cm     = last["date"]
    cdim   = days_in_month(cm.year, cm.month)
    cdr    = last["demand"] / cdim
    dr_map = {}
    for d in range(TODAY.day, cdim + 1):
        dr_map[date(cm.year, cm.month, d)] = cdr
    for _, row in fore_df.iterrows():
        fd  = row["date"]; dim = days_in_month(fd.year, fd.month); dr = row["forecast"] / dim
        for d in range(1, dim + 1):
            dr_map[date(fd.year, fd.month, d)] = dr

    pending    = []
    daily_rows = []
    orders     = []
    mop = {}; mcl = {}; mrv = {}; mord = {}; mdr = {}

    for d in sorted(dr_map.keys()):
        dr  = dr_map[d]
        mk  = date(d.year, d.month, 1)

        # Receive deliveries due today
        recv    = sum(q for (dd, q) in pending if dd == d)
        pending = [(dd, q) for (dd, q) in pending if dd != d]
        stock  += recv

        if mk not in mop:
            mop[mk]  = stock; mrv[mk] = 0
            mord[mk] = [];    mdr[mk] = dr
        mrv[mk] += recv; mdr[mk] = dr

        # Trigger = ROP + lead_time × daily_rate
        trig   = rop + lead * dr
        placed = False; oqty = 0

        if stock <= trig and not pending:
            oqty    = round_up_to(max(COVER_DAYS * dr, min_order), ROUND_TO)
            dd      = d + timedelta(days=lead)
            pending.append((dd, oqty))
            orders.append({"date": d, "qty": oqty, "delivery": dd})
            mord[mk].append({"date": d, "qty": oqty, "delivery": dd})
            placed = True

        stock = max(stock - dr, 0)
        mcl[mk] = stock

        daily_rows.append({
            "date":   d,
            "month":  d.strftime("%b %Y"),
            "day":    d.day,
            "dr":     round(dr, 1),
            "stock":  round(stock, 0),
            "rop":    rop,
            "trig":   round(trig, 0),
            "recv":   round(recv, 0),
            "placed": placed,
            "oqty":   oqty,
        })

    # Build monthly summary
    monthly = []
    for mk in sorted(mop.keys()):
        dr_m = mdr.get(mk, 0); dim = days_in_month(mk.year, mk.month)
        is_cur = (mk == date(TODAY.year, TODAY.month, 1))
        fr     = fore_df[fore_df["date"] == mk]
        fc_val = int(fr.iloc[0]["forecast"]) if not fr.empty else int(dr_m * dim)
        ords   = mord.get(mk, [])
        mo2    = mop.get(mk, 0)
        dtr    = (mo2 - rop) / dr_m if dr_m > 0 else float("inf")
        proj_rop = (mk + timedelta(days=max(0, int(dtr)))).strftime("%Y-%m-%d") if dtr >= 0 else "PAST"
        monthly.append({
            "month":    mk.strftime("%b %Y") + (" ▶" if is_cur else ""),
            "mkey":     mk,
            "dr":       round(dr_m, 1),
            "dim":      dim,
            "forecast": fc_val,
            "opening":  int(mop.get(mk, 0)),
            "received": int(mrv.get(mk, 0)),
            "closing":  int(mcl.get(mk, 0)),
            "rop":      int(rop),
            "trig":     round(rop + lead * dr_m, 0),
            "proj_rop": proj_rop,
            "ord_dates":" / ".join(o["date"].strftime("%Y-%m-%d") for o in ords) if ords else "—",
            "ord_qtys": " / ".join(str(o["qty"]) for o in ords) if ords else "—",
            "del_dates":" / ".join(o["delivery"].strftime("%Y-%m-%d") for o in ords) if ords else "—",
            "n_orders": len(ords),
        })
    return daily_rows, monthly, orders


# ── Matplotlib charts ─────────────────────────────────────────────────────────

def to_img(fig, dpi=120):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight", facecolor="white")
    buf.seek(0); plt.close(fig); return buf


def setup_ax(ax):
    ax.set_facecolor("white")
    ax.grid(True, color="#EEEEEE", linewidth=0.8, zorder=0)
    ax.spines[["top", "right"]].set_visible(False)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))


def chart1_demand(prod, hist_df, fore_df, fitted):
    """Chart 1: Historical demand, HW fitted values, and forecast."""
    fig, ax = plt.subplots(figsize=(12, 5))
    fig.patch.set_facecolor("white"); setup_ax(ax)

    h_labels = [r["date"].strftime("%b %Y") for _, r in hist_df.iterrows()]
    h_demand  = [int(r["demand"])            for _, r in hist_df.iterrows()]
    h_fitted  = [int(f)                      for f    in fitted]
    f_labels  = [r["date"].strftime("%b %Y") for _, r in fore_df.iterrows()]
    f_vals    = [int(r["forecast"])           for _, r in fore_df.iterrows()]

    all_labels = h_labels + f_labels
    xh = range(len(h_labels))
    xf = range(len(h_labels) - 1, len(h_labels) - 1 + len(f_labels))

    # Historical actual
    ax.bar(list(xh), h_demand, color=C_BLUE, alpha=0.25, label="Actual Demand", zorder=2, width=0.5)
    ax.plot(xh, h_demand, "o-", color=C_BLUE, lw=2, ms=7, zorder=4)

    # HW fitted
    ax.plot(xh, h_fitted, "s--", color=C_TEAL, lw=1.8, ms=6, zorder=4,
            label="HW Fitted Values", alpha=0.85)

    # Forecast
    ax.plot(xf, f_vals, "D-", color=C_ORANGE, lw=2.2, ms=8, zorder=4, label="HW Forecast")

    # Shading
    ax.axvspan(len(h_labels) - 1.5, len(all_labels) - 0.5, alpha=0.04, color=C_ORANGE)
    ax.axvline(x=len(h_labels) - 1.5, color=C_GRAY, lw=1, ls=":", zorder=1)

    ymax = max(max(h_demand, default=0), max(f_vals, default=0)) * 1.2

    # Data labels
    for xi, yi in zip(xh, h_demand):
        ax.annotate(f"{yi:,}", (xi, yi), xytext=(0, 8), textcoords="offset points",
                    ha="center", fontsize=7.5, color=C_BLUE, fontweight="bold")
    for xi, yi in zip(xf, f_vals):
        ax.annotate(f"{yi:,}", (xi, yi), xytext=(0, 8), textcoords="offset points",
                    ha="center", fontsize=7.5, color=C_ORANGE, fontweight="bold")

    ax.text(len(h_labels) - 1.35, ymax * 0.97,
            "◀ Historical         Forecast ▶",
            fontsize=8.5, color=C_GRAY, va="top", ha="center")

    ax.set_xticks(range(len(all_labels)))
    ax.set_xticklabels(all_labels, rotation=40, ha="right", fontsize=8.5)
    ax.set_title(f"Chart 1  —  Product {prod}: Demand Forecast (Holt-Winters Multiplicative)",
                 fontsize=12, fontweight="bold", color=C_BLUE, pad=10)
    ax.set_ylabel("Demand (Units)", fontsize=9)
    ax.set_xlabel("Month", fontsize=9)
    ax.legend(loc="upper left", framealpha=0.9, fontsize=9)
    ax.set_ylim(0, ymax)
    fig.tight_layout(pad=1.2)
    return to_img(fig)


def chart2_stock(prod, monthly, orders):
    """Chart 2: Monthly stock levels vs ROP and trigger."""
    fig, ax = plt.subplots(figsize=(12, 5.5))
    fig.patch.set_facecolor("white"); setup_ax(ax)

    labels = [m["month"].replace(" ▶", "") for m in monthly]
    opens  = [m["opening"]  for m in monthly]
    closes = [m["closing"]  for m in monthly]
    rops   = [m["rop"]      for m in monthly]
    trigs  = [m["trig"]     for m in monthly]
    x      = range(len(labels))

    ax.fill_between(x, opens, closes, alpha=0.10, color=C_BLUE)
    ax.plot(x, opens,  "o-", color=C_BLUE,   lw=2.2, ms=8, label="Opening Stock",  zorder=4)
    ax.plot(x, closes, "s-", color=C_GREEN,  lw=2.2, ms=8, label="Closing Stock",  zorder=4)
    ax.plot(x, rops,   "--", color=C_RED,    lw=1.8,        label="Reorder Point (ROP)", zorder=3)
    ax.plot(x, trigs,  ":",  color=C_PURPLE, lw=1.6,        label="Order Trigger Level", zorder=3)

    # Data labels
    for xi, yi in zip(x, opens):
        ax.annotate(f"{yi:,}", (xi, yi), xytext=(0, 9), textcoords="offset points",
                    ha="center", fontsize=7.5, color=C_BLUE, fontweight="bold")
    for xi, yi in zip(x, closes):
        ax.annotate(f"{yi:,}", (xi, yi), xytext=(0, -15), textcoords="offset points",
                    ha="center", fontsize=7.5, color=C_GREEN)

    # Order event badges
    ord_map = {o["date"].strftime("%b %Y"): o["qty"] for o in orders}
    for xi, lbl in enumerate(labels):
        if lbl in ord_map:
            ax.annotate(f"ORDER\n{ord_map[lbl]:,}u",
                        xy=(xi, trigs[xi]), xytext=(0, -36),
                        textcoords="offset points",
                        fontsize=7.5, color=C_RED, fontweight="bold", ha="center",
                        bbox=dict(boxstyle="round,pad=0.25", fc="#FCE4D6", ec=C_RED, alpha=0.95),
                        arrowprops=dict(arrowstyle="-|>", color=C_RED, lw=1.2))

    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=40, ha="right", fontsize=8.5)
    ax.set_title(f"Chart 2  —  Product {prod}: Monthly Stock Level vs Reorder Point & Trigger",
                 fontsize=12, fontweight="bold", color=C_BLUE, pad=10)
    ax.set_ylabel("Stock Level (Units)", fontsize=9)
    ax.set_xlabel("Month", fontsize=9)
    ax.legend(loc="upper right", framealpha=0.9, fontsize=9)
    fig.tight_layout(pad=1.2)
    return to_img(fig)


def chart3_daily(prod, daily_rows, rop):
    """Chart 3: Day-by-day stock simulation with order/delivery markers."""
    fig, ax = plt.subplots(figsize=(14, 5.5))
    fig.patch.set_facecolor("white"); setup_ax(ax)

    dates  = [r["date"]  for r in daily_rows]
    stocks = [r["stock"] for r in daily_rows]
    trigs  = [r["trig"]  for r in daily_rows]

    ax.fill_between(dates, stocks, alpha=0.12, color=C_BLUE)
    ax.plot(dates, stocks, "-", color=C_BLUE,   lw=1.6, label="Daily Stock Level", zorder=3)
    ax.axhline(y=rop, color=C_RED,    lw=1.5, ls="--", label=f"Reorder Point ({rop:,})", zorder=2)
    ax.plot(dates, trigs,  ":",  color=C_PURPLE, lw=1.1,
            label="Order Trigger", alpha=0.7, zorder=2)

    # Mark order placements and receipts
    op = [r for r in daily_rows if r["placed"]]
    rv = [r for r in daily_rows if r["recv"] > 0]
    if op:
        ax.scatter([r["date"] for r in op], [r["stock"] for r in op],
                   color=C_RED, s=80, marker="v", zorder=5, label="Order Placed",
                   linewidths=0.5, edgecolors="white")
    if rv:
        ax.scatter([r["date"] for r in rv], [r["stock"] for r in rv],
                   color=C_GREEN, s=80, marker="^", zorder=5, label="Stock Received",
                   linewidths=0.5, edgecolors="white")

    # Month boundaries
    seen = set()
    for r in daily_rows:
        mk = (r["date"].year, r["date"].month)
        if mk not in seen:
            seen.add(mk)
            if r["date"] > TODAY:
                ax.axvline(x=r["date"], color="#DDDDDD", lw=0.8, zorder=1)
                ax.text(r["date"], max(stocks) * 1.01,
                        r["date"].strftime("%b"), fontsize=7.5, color="#AAAAAA",
                        ha="left", va="bottom")

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
    ax.xaxis.set_major_locator(mdates.MonthLocator())
    fig.autofmt_xdate(rotation=40, ha="right")
    ax.set_title(
        f"Chart 3  —  Product {prod}: Day-by-Day Simulation  "
        f"(▼ = order placed  ▲ = stock received)",
        fontsize=12, fontweight="bold", color=C_BLUE, pad=10)
    ax.set_ylabel("Stock Level (Units)", fontsize=9)
    ax.set_xlabel("Date", fontsize=9)
    ax.legend(loc="upper right", framealpha=0.9, fontsize=9)
    fig.tight_layout(pad=1.2)
    return to_img(fig)


# ── Excel result sheet ────────────────────────────────────────────────────────

def write_sheet(wb, prod, hist_df, fore_df, L, B, S, fitted,
                daily_rows, monthly, orders, cfg, fore_cfg, img1, img2, img3):
    sname = f"Result — {prod}"
    if sname in wb.sheetnames: del wb[sname]
    ws = wb.create_sheet(sname)

    lead  = int(fore_cfg.get("Lead Time (days)", 20))
    alpha = fore_cfg.get("Alpha - Level",   0.3)
    beta  = fore_cfg.get("Beta - Trend",    0.1)
    gamma = fore_cfg.get("Gamma - Seasonal",0.2)
    phi   = fore_cfg.get("Damping Factor",  0.88)
    rop   = cfg["rop"]; inv = cfg["inventory"]; mo_min = cfg["min_order"]
    nm    = len(monthly)

    # ── Title ─────────────────────────────────────────────────────────────────
    t = ws.cell(row=1, column=1,
                value=f"Inventory Forecast & Procurement Plan  —  Product: {prod}")
    t.font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    ws.merge_cells("A1:N1"); ws.row_dimensions[1].height = 28

    # ── KPI bar ───────────────────────────────────────────────────────────────
    for ki, (lbl, val) in enumerate([
        ("Current Stock", inv), ("Reorder Point", rop),
        ("Min Order",     mo_min), ("Cover Days", COVER_DAYS), ("Lead Time", lead)
    ]):
        col = ki * 2 + 1
        lc  = ws.cell(row=2, column=col, value=lbl)
        lc.font = Font(name="Arial", size=9, color="595959")
        lc.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        vc  = ws.cell(row=3, column=col, value=val)
        vc.font = Font(name="Arial", size=12, bold=True, color="1F4E79")
        vc.fill = YLW_FILL
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = THIN
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 24

    # ── Procurement plan table ────────────────────────────────────────────────
    sec(ws, 5, 1, "PROCUREMENT PLAN — Monthly Summary", 14)
    HDRS   = ["Month", "Days", "Forecast", "Daily Rate", "Trigger Level",
              "Opening Stock", "Received", "Closing Stock", "ROP",
              "Date Hits ROP", "Order Date(s)", "Order Qty",
              "Delivery Date(s)", "Status"]
    WIDTHS = [14, 8, 12, 10, 13, 13, 11, 13, 10, 14, 16, 12, 16, 20]
    HR = 6
    for i, (h, w) in enumerate(zip(HDRS, WIDTHS), 1): hcell(ws, HR, i, h, w)
    ws.row_dimensions[HR].height = 28
    # no freeze

    for i, r in enumerate(monthly, HR+1):
        is_o = r["n_orders"] > 0
        rf   = RED_FILL if is_o else (ALT_FILL if i % 2 == 0 else WH_FILL)
        oc   = "C00000" if is_o else "000000"
        sc   = "C00000" if is_o else "375623"
        stat = "⚠  Order placed" if is_o else "✓  OK — sufficient stock"
        dcell(ws,i, 1, r["month"],    bold=True,      fill=rf)
        dcell(ws,i, 2, r["dim"],                      fill=rf)
        dcell(ws,i, 3, r["forecast"],                 fill=rf)
        dcell(ws,i, 4, r["dr"],                       fill=rf)
        dcell(ws,i, 5, r["trig"],                     fill=rf)
        dcell(ws,i, 6, r["opening"],                  fill=rf)
        dcell(ws,i, 7, r["received"] if r["received"] > 0 else "—",
              color="375623" if r["received"] > 0 else "000000", fill=rf)
        dcell(ws,i, 8, r["closing"],                  fill=rf)
        dcell(ws,i, 9, r["rop"],                      fill=rf)
        dcell(ws,i,10, r["proj_rop"],                 fill=rf)
        dcell(ws,i,11, r["ord_dates"], bold=is_o, color=oc, fill=rf)
        dcell(ws,i,12, r["ord_qtys"],  bold=is_o, color=oc, fill=rf)
        dcell(ws,i,13, r["del_dates"],                fill=rf)
        dcell(ws,i,14, stat, bold=is_o, color=sc,     fill=rf)
        ws.row_dimensions[i].height = 18

    # ── Calculation tables ────────────────────────────────────────────────────
    CR = HR + nm + 3

    # A. Settings
    sec(ws, CR, 1, "A.  SETTINGS USED", 8, HDR_FILL); r = CR + 1
    for lbl, val, note in [
        ("Current Inventory",      inv,       "Units on hand today"),
        ("Reorder Point (ROP)",    rop,       "Delivery arrives when stock = ROP (you set this)"),
        ("Minimum Order",          mo_min,    "Quantity floor only — does NOT affect order timing"),
        ("Lead Time",              lead,      "Days from order placement to delivery"),
        ("Cover Days",             COVER_DAYS,"Each order covers this many days of demand"),
        ("Alpha  (level)",         alpha,     "Holt-Winters: level smoothing weight [0,1]"),
        ("Beta   (trend)",         beta,      "Holt-Winters: trend smoothing weight [0,1]"),
        ("Gamma  (seasonal)",      gamma,     "Holt-Winters: seasonal update weight [0,1]"),
        ("Phi    (damping)",       phi,       "Holt-Winters: trend damping factor [0.8–1.0]"),
    ]:
        f = GRN_FILL if r % 2 == 0 else WH_FILL
        dcell(ws, r, 1, lbl, bold=True, fill=f)
        dcell(ws, r, 2, val, color="1F4E79", fill=f)
        c = ws.cell(row=r, column=3, value=note)
        c.font = Font(name="Arial", size=10, color="595959")
        c.fill = f; c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = THIN
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 17; r += 1

    # B. Formula
    r += 1; sec(ws, r, 1, "B.  ORDER TRIGGER FORMULA & HOLT-WINTERS EXPLANATION", 8, HDR_FILL); r += 1
    for lbl, txt in [
        ("HW Method",
         "Holt-Winters Multiplicative: L[t] = α(y[t]/S[t]) + (1-α)(L[t-1]+φB[t-1])  |  "
         "B[t] = β(L[t]-L[t-1]) + (1-β)φB[t-1]  |  S[t] = γ(y[t]/L[t]) + (1-γ)S[t]"),
        ("Init",
         "Seasonal factors S initialized from col E multipliers, then updated each month with gamma."),
        ("Forecast",
         f"F[t+h] = (L[t] + damped_trend[h]) × future_multiplier[h]   "
         f"where damped_trend[h] = B[t] × Σφⁱ for i=1..h  (phi={phi})"),
        ("Trigger",
         "Order Trigger = ROP + Lead_Time × Daily_Rate.  "
         "Place order when stock ≤ Trigger → delivery arrives lead_time days later at stock = ROP."),
        ("Order Qty",
         f"round_up_to_{ROUND_TO}( max( {COVER_DAYS} × daily_rate,  min_order ) )"),
        ("Rule",
         "ONE pending order at a time. No new order while delivery is in transit."),
    ]:
        f  = ALT_FILL if r % 2 == 0 else WH_FILL
        c1 = ws.cell(row=r, column=1, value=lbl)
        c1.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        c1.fill = f; c1.border = THIN
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2 = ws.cell(row=r, column=2, value=txt)
        c2.font = Font(name="Arial", size=10)
        c2.fill = f; c2.border = THIN
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 22; r += 1

    # C. Historical data + HW smoothing output
    r += 1; sec(ws, r, 1, "C.  HISTORICAL DATA & HOLT-WINTERS SMOOTHING", 9, HDR_FILL); r += 1
    for i, h in enumerate(["Month", "Days", "Demand", "Seasonal Index\n(col E, init)",
                            "HW Level (L)", "HW Trend (B)", "HW Seasonal (S)",
                            "HW Fitted", "Error %"], 1):
        hcell(ws, r, i, h, w=[12,7,12,16,14,13,15,12,10][i-1], fill=HDR2_FILL)
    ws.row_dimensions[r].height = 28; r += 1

    for j, (_, hrow) in enumerate(hist_df.iterrows()):
        dim  = days_in_month(hrow["date"].year, hrow["date"].month)
        dem  = int(hrow["demand"])
        fit  = int(fitted[j])
        err  = round((fit - dem) / dem * 100, 1) if dem != 0 else 0
        err_color = "C00000" if abs(err) > 20 else ("BA6200" if abs(err) > 10 else "375623")
        dcell(ws, r, 1, hrow["date"].strftime("%b %Y"))
        dcell(ws, r, 2, dim)
        dcell(ws, r, 3, dem)
        dcell(ws, r, 4, round(hrow["multiplier"], 2))
        dcell(ws, r, 5, round(float(L[j]), 1),  color="1F4E79")
        dcell(ws, r, 6, round(float(B[j]), 2),  color="375623")
        dcell(ws, r, 7, round(float(S[j]), 3),  color="7030A0")
        dcell(ws, r, 8, fit,                     color="1F4E79", bold=True)
        dcell(ws, r, 9, f"{err:+.1f}%",          color=err_color)
        ws.row_dimensions[r].height = 17; r += 1

    # D. Forecast results
    r += 1; sec(ws, r, 1, "D.  DEMAND FORECAST RESULTS (Holt-Winters)", 9, HDR_FILL); r += 1
    for i, h in enumerate(["Month", "Days", "Future\nMultiplier",
                            "De-trended\nLevel Proj.", "Final Forecast\n(× multiplier)",
                            "Daily Rate", "Lead Demand\n(lead×daily)", "Trigger Level",
                            "Order Qty if\nTriggered"], 1):
        hcell(ws, r, i, h, w=[12,7,14,16,18,11,15,14,16][i-1], fill=HDR2_FILL)
    ws.row_dimensions[r].height = 30; r += 1

    phi_acc = 0.0
    for fi, (_, fr) in enumerate(fore_df.iterrows()):
        fd       = fr["date"]
        dim      = days_in_month(fd.year, fd.month)
        dr       = fr["forecast"] / dim
        phi_acc += phi ** (fi + 1)
        level_proj = round(float(L[-1]) + phi_acc * float(B[-1]), 1)
        lead_dem = round(lead * dr, 0)
        trig     = round(rop + lead * dr, 0)
        oqty     = round_up_to(max(COVER_DAYS * dr, cfg["min_order"]), ROUND_TO)
        dcell(ws, r, 1, fd.strftime("%b %Y"))
        dcell(ws, r, 2, dim)
        dcell(ws, r, 3, round(fr["multiplier"], 2))
        dcell(ws, r, 4, level_proj)
        dcell(ws, r, 5, int(fr["forecast"]), color="1F4E79", bold=True)
        dcell(ws, r, 6, round(dr, 1))
        dcell(ws, r, 7, lead_dem)
        dcell(ws, r, 8, trig)
        dcell(ws, r, 9, oqty)
        ws.row_dimensions[r].height = 17; r += 1

    # ── Daily table ───────────────────────────────────────────────────────────
    r += 2
    note = ws.cell(row=r, column=1,
                   value="Day-by-Day Simulation  —  Filter by Month (▼) to zoom into one month")
    note.font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.merge_cells(f"A{r}:I{r}"); r += 1; DR = r

    for i, (h, w) in enumerate(zip(
        ["Date","Month","Day","Daily Rate","Stock Level",
         "ROP","Trigger Level","Order Placed?","Order Qty"],
        [13,13,6,11,13,10,14,13,11]), 1):
        hcell(ws, DR, i, h, w, fill=HDR2_FILL)
    ws.row_dimensions[DR].height = 22

    for ri, d in enumerate(daily_rows, DR+1):
        rf = RED_FILL if d["placed"] else (ALT_FILL if ri % 2 == 0 else WH_FILL)
        for ci, val in enumerate([
            d["date"].strftime("%Y-%m-%d"), d["month"], d["day"], d["dr"],
            int(d["stock"]), int(d["rop"]), int(d["trig"]),
            "✓ Yes" if d["placed"] else "",
            d["oqty"] if d["oqty"] > 0 else ""
        ], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=9,
                          bold=d["placed"] and ci in (1, 8, 9),
                          color="C00000" if d["placed"] and ci in (8, 9) else "000000")
            c.fill = rf
            c.alignment = Alignment(horizontal="center"); c.border = THIN
        ws.row_dimensions[ri].height = 14

    if daily_rows:
        tbl_end = DR + len(daily_rows)
        tbl = Table(displayName=f"Daily_{prod.replace(' ','_')}",
                    ref=f"A{DR}:I{tbl_end}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl)

    # ── Embed 3 charts in column K ─────────────────────────────────────────────
    for col_l in ["K","L","M","N","O","P","Q","R","S","T","U","V","W"]:
        ws.column_dimensions[col_l].width = 11.5

    for img_buf, anchor in [(img1,"K1"), (img2,"K25"), (img3,"K49")]:
        img = XLImage(img_buf)
        img.width = 820; img.height = 380
        ws.add_image(img, anchor)

    for ri in range(1, 80):
        if ws.row_dimensions[ri].height is None or ws.row_dimensions[ri].height < 15:
            ws.row_dimensions[ri].height = 16


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = load_workbook(INPUT_FILE)
    prod_cfg, fore_cfg = read_settings(wb)

    alpha = fore_cfg.get("Alpha - Level",    0.3)
    beta  = fore_cfg.get("Beta - Trend",     0.1)
    gamma = fore_cfg.get("Gamma - Seasonal", 0.2)
    phi   = fore_cfg.get("Damping Factor",   0.88)
    lead  = int(fore_cfg.get("Lead Time (days)", 20))

    products = [s for s in wb.sheetnames
                if s not in ["Settings"] and not s.startswith("Result")]
    print(f"Products: {products}  |  Today: {TODAY}")
    print(f"Method: Holt-Winters Multiplicative")
    print(f"alpha={alpha}  beta={beta}  gamma={gamma}  phi={phi}  lead={lead}\n")

    for prod in products:
        print(f"─── {prod} ───")
        cfg = prod_cfg.get(prod, {"inventory":0, "rop":0, "min_order":500})

        hist_df, fut_df        = read_sheet(wb[prod])
        fore_df, L, B, S, fit = run_forecast(hist_df, fut_df, alpha, beta, gamma, phi)
        daily_rows, monthly, orders = simulate(hist_df, fore_df, cfg, lead)

        print(f"  Generating 3 charts...")
        img1 = chart1_demand(prod, hist_df, fore_df, fit)
        img2 = chart2_stock(prod, monthly, orders)
        img3 = chart3_daily(prod, daily_rows, int(cfg["rop"]))

        write_sheet(wb, prod, hist_df, fore_df, L, B, S, fit,
                    daily_rows, monthly, orders, cfg, fore_cfg, img1, img2, img3)

        print(f"  {'Month':<16}{'Open':>8}{'Trig':>8}{'Close':>8}  Orders")
        for mr in monthly:
            print(f"  {mr['month']:<16}{mr['opening']:>8}"
                  f"{mr['trig']:>8.0f}{mr['closing']:>8}  "
                  f"{mr['ord_dates']}  qty:{mr['ord_qtys']}")
        print()

    # Add Gamma setting to Settings sheet if not present
    ws_s = wb["Settings"]
    has_gamma = any("gamma" in str(r[0] or "").lower()
                    for r in ws_s.iter_rows(values_only=True) if r[0])
    if not has_gamma:
        for row in ws_s.iter_rows():
            if row[0].value and "beta" in str(row[0].value).lower():
                ins_row = row[0].row + 1
                ws_s.insert_rows(ins_row)
                ws_s.cell(row=ins_row, column=1, value="Gamma - Seasonal").font = \
                    Font(name="Arial", size=10, bold=True)
                c = ws_s.cell(row=ins_row, column=2, value=0.2)
                c.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
                c.fill = YLW_FILL
                ws_s.cell(row=ins_row, column=4, value=
                    "Holt-Winters seasonal smoothing [0.1=stable … 0.5=adaptive]").font = \
                    Font(name="Arial", size=10, color="595959", italic=True)
                print(f"  Added 'Gamma - Seasonal' to Settings sheet at row {ins_row}")
                break

    wb.save(INPUT_FILE)
    print(f"Saved → {INPUT_FILE}")
    print("6 charts (3 per product) embedded as PNG images.")


if __name__ == "__main__":
    main()
